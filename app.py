import streamlit as st
import pandas as pd
import numpy as np
import io
from statsmodels.tsa.statespace.sarimax import SARIMAX
import warnings
warnings.filterwarnings('ignore')

# Try to import xlsxwriter, fall back to openpyxl if not available
try:
    import xlsxwriter
    EXCEL_ENGINE = "xlsxwriter"
    XLSXWRITER_AVAILABLE = True
except ImportError:
    EXCEL_ENGINE = "openpyxl"
    XLSXWRITER_AVAILABLE = False
    st.warning("⚠️ xlsxwriter not available. Using openpyxl for Excel export (limited formatting).")

st.set_page_config(layout="wide", page_title="Spare Parts Forecast", page_icon="📈")
st.title("📈 Spare Parts Forecast (MH-Family SARIMA with Log Transform)")

# —————————————————————————————
# 1) File Uploaders
# —————————————————————————————

st.write(
    """
    **Step 1 – Upload your data files**  
    - **Parts × Family × Month × Sales:**  
      Columns: `Part`, `Family`, `Month`, `Sales`.  
      (e.g. "Fork_001", "Linde-Forks", "2023-05-01", 12)  
    - **FleetHours (optional):**  
      Columns: `Month`, `FleetHours` (total MH-fleet runtime in Saudi Arabia).  
      This will be used as an exogenous regressor (usage often drops during Ramadan/Eid).
    """
)

uploaded_parts = st.file_uploader("Upload Parts×Family×Month×Sales file", type=["xlsx", "csv"])
uploaded_exog  = st.file_uploader("Upload FleetHours file (Month, FleetHours) (Optional)", type=["xlsx", "csv"])

if not uploaded_parts:
    st.info("📥 Please upload the Parts×Family×Month×Sales file to proceed.")
    st.stop()

# —————————————————————————————
# 2) Load & Validate Input Data
# —————————————————————————————

@st.cache_data
def load_parts_df(f):
    try:
        if str(f.name).lower().endswith(".csv"):
            return pd.read_csv(f)
        else:
            return pd.read_excel(f)
    except Exception as e:
        st.error(f"❌ Error loading file: {str(e)}")
        st.stop()

@st.cache_data
def load_exog_df(f):
    try:
        if str(f.name).lower().endswith(".csv"):
            return pd.read_csv(f)
        else:
            return pd.read_excel(f)
    except Exception as e:
        st.error(f"❌ Error loading exogenous file: {str(e)}")
        st.stop()

df = load_parts_df(uploaded_parts)

# Validate required columns
required_cols = {"Part", "Family", "Month", "Sales"}
if not required_cols.issubset(df.columns):
    st.error(f"❌ Your file must contain columns: {required_cols}. Found: {list(df.columns)}")
    st.stop()

# Data cleaning and validation
original_shape = df.shape
df = df.dropna(subset=["Part", "Family", "Month", "Sales"]).copy()

if df.empty:
    st.error("❌ No valid data found after removing rows with missing values.")
    st.stop()

df["Part"]   = df["Part"].astype(str).str.strip()
df["Family"] = df["Family"].astype(str).str.strip()

# Parse Month → datetime (coerce invalid)
df["Month"] = pd.to_datetime(df["Month"], errors="coerce")
n_invalid = df["Month"].isna().sum()
if n_invalid > 0:
    st.warning(f"⚠️ Dropping {n_invalid} rows with invalid Month parse.")
    df = df.dropna(subset=["Month"])

if df.empty:
    st.error("❌ No valid data found after date parsing.")
    st.stop()

# Convert sales to numeric and handle negatives
df["Sales"] = pd.to_numeric(df["Sales"], errors="coerce").fillna(0)
negative_sales = (df["Sales"] < 0).sum()
if negative_sales > 0:
    st.warning(f"⚠️ Found {negative_sales} negative sales values. Converting to 0.")
    df["Sales"] = df["Sales"].clip(lower=0)

df["Sales"] = df["Sales"].astype(float)

# Performance warnings
if df.shape[0] > 10000:
    st.warning("⚠️ Large dataset detected. Processing may take several minutes.")

unique_parts = df["Part"].nunique()
if unique_parts > 1000:
    st.info("💡 Consider processing in smaller batches for better performance.")

st.write(f"Data shape: {original_shape} → {df.shape} (after cleaning)")
st.write(f"Unique parts: {unique_parts}, Unique families: {df['Family'].nunique()}")
st.write("Preview of uploaded data:")
st.dataframe(df.head())

# —————————————————————————————
# 3) Load Optional Exogenous (FleetHours) + Build Ramadan Flag
# —————————————————————————————

if uploaded_exog:
    exog_df = load_exog_df(uploaded_exog)
    if not {"Month", "FleetHours"}.issubset(exog_df.columns):
        st.error("❌ Exogenous file must have columns: Month, FleetHours.")
        st.stop()
    
    exog_df["Month"] = pd.to_datetime(exog_df["Month"], errors="coerce")
    ninv = exog_df["Month"].isna().sum()
    if ninv > 0:
        st.warning(f"⚠️ Dropping {ninv} rows from exogenous data with invalid Month.")
        exog_df = exog_df.dropna(subset=["Month"])
    
    exog_df["FleetHours"] = pd.to_numeric(exog_df["FleetHours"], errors="coerce").fillna(0)
    
    # Handle negative fleet hours
    negative_fleet = (exog_df["FleetHours"] < 0).sum()
    if negative_fleet > 0:
        st.warning(f"⚠️ Found {negative_fleet} negative FleetHours values. Converting to 0.")
        exog_df["FleetHours"] = exog_df["FleetHours"].clip(lower=0)
    
    # Reindex exog to Month→FleetHours
    fleet_hours = (
        exog_df
        .groupby(pd.Grouper(key="Month", freq="MS"))["FleetHours"]
        .sum()
        .sort_index()
    )
else:
    # If no exog provided, create a zero series placeholder
    fleet_hours = pd.Series(dtype=float)

# Build a robust monthly Ramadan indicator (Saudi Arabia)
def build_ramadan_flag(idx):
    """
    Build Ramadan indicator for Saudi Arabia.
    Returns 1 for months that overlap with Ramadan, 0 otherwise.
    """
    # Ramadan months by year (approximate, based on lunar calendar)
    ramadan_months = {
        2022: [4, 5],      # Apr-May 2022
        2023: [3, 4],      # Mar-Apr 2023  
        2024: [3, 4],      # Mar-Apr 2024
        2025: [2, 3],      # Feb-Mar 2025
        2026: [2, 3],      # Feb-Mar 2026
        2027: [1, 2],      # Jan-Feb 2027
        2028: [1, 12],     # Jan + Dec 2028 (wraps around)
        2029: [12],        # Dec 2029
        2030: [11, 12],    # Nov-Dec 2030
    }
    
    flags = []
    for ts in idx:
        y, m = ts.year, ts.month
        if y in ramadan_months and m in ramadan_months[y]:
            flags.append(1)
        else:
            flags.append(0)
    return pd.Series(flags, index=idx)

# —————————————————————————————
# 4) Aggregate (Family, Month) at both Family and Part Levels
# —————————————————————————————

# 4.a Family-level: for parameter selection
family_ts = (
    df
    .groupby([pd.Grouper(key="Month", freq="MS"), "Family"])["Sales"]
    .sum()
    .unstack(fill_value=0)
    .sort_index()
)

# 4.b Part-level: sum any duplicates
part_ts_df = (
    df
    .groupby([pd.Grouper(key="Month", freq="MS"), "Part", "Family"])["Sales"]
    .sum()
    .reset_index()
)

# —————————————————————————————
# 5) Determine Overall Date Range & Build Month Indices
# —————————————————————————————

min_date = part_ts_df["Month"].min().replace(day=1)
max_date = part_ts_df["Month"].max().replace(day=1)

historical_months = pd.date_range(start=min_date, end=max_date, freq="MS")
forecast_start = max_date + pd.DateOffset(months=1)
forecast_end   = forecast_start + pd.DateOffset(months=11)
forecast_months = pd.date_range(start=forecast_start, end=forecast_end, freq="MS")

st.write(
    f"📅 Historical data: {min_date.strftime('%Y-%m')} → {max_date.strftime('%Y-%m')} "
    f"({len(historical_months)} months)."
)
st.write(
    f"🔮 Forecast horizon: {forecast_start.strftime('%Y-%m')} → {forecast_end.strftime('%Y-%m')}."
)

# Validate sufficient history
if len(historical_months) < 12:
    st.warning("⚠️ Less than 12 months of historical data. Forecasts may be less reliable.")

# Build seasonal Ramadan flags for historical + forecast months
all_months = historical_months.union(forecast_months)
ramadan_flag = build_ramadan_flag(all_months)

# Build a complete exogenous DataFrame indexed by Month: columns: FleetHours, IsRamadan
exog_full = pd.DataFrame(index=all_months)
# 1) Populate FleetHours: reindex to all_months (fill missing with median or 0)
if not fleet_hours.empty:
    median_fleet = fleet_hours.median() if fleet_hours.median() > 0 else 0
    exog_full["FleetHours"] = fleet_hours.reindex(all_months, fill_value=median_fleet)
else:
    exog_full["FleetHours"] = 0

# 2) Populate Ramadan indicator
exog_full["IsRamadanMonth"] = ramadan_flag

st.write("Here is a preview of the exogenous DataFrame (FleetHours + Ramadan flag):")
st.dataframe(exog_full.head(12))

# —————————————————————————————
# 6) Pre-Fit Family SARIMA via Grid Search (optimized)
# —————————————————————————————

st.write("### ▶️ Pre-fitting SARIMA on each Family (this may take a moment)…")

# Optimized parameter ranges to reduce computation time
ps = ds = qs = range(0, 2)      # test p,d,q ∈ {0,1}
Ps = Ds = Qs = range(0, 2)      # test P,D,Q ∈ {0,1}
s = 12                          # seasonal period = 12 months

family_params = {}  # { family: {"order": (p,d,q), "seasonal": (P,D,Q,12)} }
max_iterations_per_family = 32  # Limit grid search iterations (2^5 combinations)

family_progress = st.progress(0.0)
family_status = st.empty()

families = list(family_ts.columns)
for fam_idx, fam in enumerate(families):
    family_progress.progress((fam_idx + 1) / len(families))
    family_status.text(f"Fitting family {fam_idx + 1}/{len(families)}: {fam}")
    
    ts_fam = family_ts[fam].reindex(historical_months, fill_value=0)
    
    # If family has almost no data, use conservative default
    if ts_fam.sum() < 1 or (ts_fam > 0).sum() < 3:
        family_params[fam] = {
            "order":    (0, 1, 1),
            "seasonal": (0, 1, 1, 12),
        }
        continue

    best_aic = np.inf
    best_cfg = None
    iteration_count = 0

    # Build exogenous matrix for in-sample: 2 columns [FleetHours, IsRamadanMonth]
    exog_insample = exog_full.loc[historical_months, ["FleetHours", "IsRamadanMonth"]].values

    for p in ps:
        for d in ds:
            for q in qs:
                for P in Ps:
                    for D in Ds:
                        for Q in Qs:
                            iteration_count += 1
                            if iteration_count > max_iterations_per_family:
                                break
                                
                            order = (p, d, q)
                            seasonal_order = (P, D, Q, s)
                            try:
                                model = SARIMAX(
                                    ts_fam,
                                    exog=exog_insample,
                                    order=order,
                                    seasonal_order=seasonal_order,
                                    enforce_stationarity=False,
                                    enforce_invertibility=False,
                                ).fit(disp=False, maxiter=50)  # Limit iterations
                                
                                if model.aic < best_aic:
                                    best_aic = model.aic
                                    best_cfg = (order, seasonal_order)
                            except Exception:
                                continue
                        if iteration_count > max_iterations_per_family:
                            break
                    if iteration_count > max_iterations_per_family:
                        break
                if iteration_count > max_iterations_per_family:
                    break
            if iteration_count > max_iterations_per_family:
                break
        if iteration_count > max_iterations_per_family:
            break

    if best_cfg is not None:
        family_params[fam] = {
            "order":    best_cfg[0],
            "seasonal": best_cfg[1],
        }
    else:
        # Conservative fallback to avoid overfitting
        family_params[fam] = {
            "order":    (0, 1, 1),
            "seasonal": (0, 1, 1, 12),
        }

family_progress.empty()
family_status.empty()

st.success("✅ Finished pre-fitting family-level SARIMA parameters.")
st.write("Family parameters discovered:")
fam_summary = pd.DataFrame.from_dict(
    {fam: {"order": str(vals["order"]), "seasonal": str(vals["seasonal"])}
     for fam, vals in family_params.items()},
    orient="index"
)
st.dataframe(fam_summary)

# —————————————————————————————
# 7) Loop Over Each Part (SKU) to Forecast
# —————————————————————————————

progress_bar = st.progress(0.0)
status_text  = st.empty()

results_data = []
parts_list = sorted(part_ts_df["Part"].unique())
total_parts = len(parts_list)

parts_with_sarima = 0
parts_with_naive  = 0
parts_failed      = 0

# Precompute: number of parts per family (for fallback allocation)
n_parts_in_family = (
    part_ts_df[["Part", "Family"]]
    .drop_duplicates()
    .groupby("Family")["Part"].nunique()
    .to_dict()
)

for idx, part in enumerate(parts_list):
    progress_bar.progress((idx + 1) / total_parts)
    status_text.text(f"Processing part {idx + 1}/{total_parts}: {part}")

    try:
        sub = part_ts_df[part_ts_df["Part"] == part][["Month", "Sales", "Family"]]
        if sub.empty:
            st.warning(f"⚠️ No data found for part {part}")
            parts_failed += 1
            continue
            
        fam = sub["Family"].iloc[0]
        raw = sub.set_index("Month")[["Sales"]].sort_index()
        raw = raw.reindex(historical_months, fill_value=0)

        # 1) Determine "active" start (first non-zero) if any
        if (raw["Sales"] > 0).any():
            start_nonzero = raw[raw["Sales"] > 0].index.min()
            ts = raw.loc[start_nonzero:, "Sales"].copy()
        else:
            ts = raw["Sales"].copy()

        non_zero_count = (ts > 0).sum()
        total_sales = ts.sum()

        # 2) If too little history or too few non-zero months, fallback to family average
        if (len(ts) < 6) or (non_zero_count < 2) or (total_sales < 1):
            fam_hist = family_ts[fam].reindex(historical_months, fill_value=0)
            last3_total = fam_hist.iloc[-3:].sum()
            n_parts = n_parts_in_family.get(fam, 1)
            
            # More robust fallback calculation
            if last3_total > 0 and n_parts > 0:
                fallback = max(0, int(round(last3_total / n_parts, 0)))
            else:
                fallback = 0
                
            fc_vals = [fallback] * 12
            method_used = "Naive (insufficient data)"
            parts_with_naive += 1

        else:
            # 3) Log-transform (log1p) instead of Box–Cox
            ts_log = np.log1p(ts)

            fam_order = family_params[fam]["order"]
            fam_seasonal = family_params[fam]["seasonal"]
            exog_insample = exog_full.loc[ts.index, ["FleetHours", "IsRamadanMonth"]].values

            # 4) Fit SARIMAX on log-transformed data
            try:
                model = SARIMAX(
                    ts_log,
                    exog=exog_insample,
                    order=fam_order,
                    seasonal_order=fam_seasonal,
                    enforce_stationarity=False,
                    enforce_invertibility=False,
                ).fit(disp=False, maxiter=100)

                # 5) Forecast next 12 months on log scale, then invert with expm1
                exog_out = exog_full.loc[forecast_months, ["FleetHours", "IsRamadanMonth"]].values
                log_forecast = model.get_forecast(steps=12, exog=exog_out).predicted_mean
                preds = np.expm1(log_forecast).round().astype(int)
                preds = np.clip(preds, 0, None)

                fc_vals = preds.tolist()
                method_used = f"SARIMA {fam_order}x{fam_seasonal[:3]}"
                parts_with_sarima += 1

            except Exception as e:
                # If SARIMAX fails, fallback to family's recent average
                fam_hist = family_ts[fam].reindex(historical_months, fill_value=0)
                recent_avg = fam_hist.iloc[-6:].mean() if len(fam_hist) >= 6 else fam_hist.mean()
                n_parts = n_parts_in_family.get(fam, 1)
                
                if recent_avg > 0 and n_parts > 0:
                    fallback = max(0, int(round(recent_avg / n_parts, 0)))
                else:
                    fallback = 0
                    
                fc_vals = [fallback] * 12
                method_used = f"Naive (SARIMA failed: {str(e)[:30]}...)"
                parts_with_naive += 1

        # 6) Build output row: historical + forecast columns
        row_data = {"Item Code": part}
        for hist_m in historical_months:
            row_data[hist_m.strftime("%b-%Y")] = int(raw.loc[hist_m, "Sales"])
        for i, fc_m in enumerate(forecast_months):
            row_data[fc_m.strftime("%b-%Y")] = int(fc_vals[i])
        row_data["Method"] = method_used
        results_data.append(row_data)

    except Exception as e:
        st.warning(f"⚠️ Skipping part {part} due to error: {str(e)}")
        parts_failed += 1
        continue

progress_bar.empty()
status_text.empty()

# —————————————————————————————
# 8) Summarize & Build Final DataFrame
# —————————————————————————————

st.write("### 📊 Processing Summary")
c1, c2, c3, c4 = st.columns(4)
c1.metric("Total Parts", total_parts)
c2.metric("SARIMA Models", parts_with_sarima)
c3.metric("Naive Forecasts", parts_with_naive)
c4.metric("Failed", parts_failed)

if not results_data:
    st.error("❌ No results generated. Please check your data.")
    st.stop()

result_df = pd.DataFrame(results_data)

# Reorder columns: Item Code → all months (historical + forecast) → Method
all_month_cols = [m.strftime("%b-%Y") for m in list(historical_months) + list(forecast_months)]
ordered_cols = ["Item Code"] + [c for c in all_month_cols if c in result_df.columns]
final_cols = ordered_cols + ["Method"]
result_df = result_df[final_cols].copy()

st.success("✅ Forecasting Completed!")
st.write(f"🔢 Parts processed: {len(result_df)}")

# Show a quick preview (last 3 historical + first 3 forecast)
preview_cols = (
    ["Item Code"]
    + all_month_cols[-3:]
    + [m.strftime("%b-%Y") for m in forecast_months[:3]]
    + ["Method"]
)
preview_cols = [c for c in preview_cols if c in result_df.columns]
st.write("**Forecast Sample (first 10 SKUs)**")
st.dataframe(result_df[preview_cols].head(10))

# Add summary statistics
st.write("### 📈 Forecast Summary Statistics")
forecast_cols = [m.strftime("%b-%Y") for m in forecast_months]
forecast_data = result_df[forecast_cols].select_dtypes(include=[np.number])

if not forecast_data.empty:
    summary_stats = pd.DataFrame({
        'Total Forecast (12 months)': forecast_data.sum(axis=1).describe(),
        'Monthly Average': forecast_data.mean(axis=1).describe()
    }).round(2)
    st.dataframe(summary_stats)

# —————————————————————————————
# 9) Download as Excel (compatible with both engines)
# —————————————————————————————

try:
    excel_df = result_df[["Item Code"] + all_month_cols].fillna("")

    # Build two-row header: month names + "Historical QTY"/"Forecasted QTY"
    header_row_1 = ["Item Code"]
    header_row_2 = [""]
    for col in all_month_cols:
        header_row_1.append(col)
        dt = pd.to_datetime(col, format="%b-%Y", errors="coerce")
        if pd.isna(dt) or (dt <= max_date):
            header_row_2.append("Historical QTY")
        else:
            header_row_2.append("Forecasted QTY")

    restructured_data = [header_row_1, header_row_2]
    for _, row in excel_df.iterrows():
        restructured_data.append(row.tolist())

    max_cols = max(len(r) for r in restructured_data)
    for r in restructured_data:
        while len(r) < max_cols:
            r.append("")

    restructured_df = pd.DataFrame(restructured_data)

    output = io.BytesIO()
    
    if XLSXWRITER_AVAILABLE:
        # Use xlsxwriter for advanced formatting
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            restructured_df.to_excel(writer, sheet_name="MH_Sales_Forecast", index=False, header=False)
            workbook = writer.book
            worksheet = writer.sheets["MH_Sales_Forecast"]

            # Enhanced formatting
            fmt_month = workbook.add_format(
                {
                    "bold": True,
                    "text_wrap": True,
                    "valign": "top",
                    "align": "center",
                    "fg_color": "#D7E4BC",
                    "border": 1,
                    "font_size": 10,
                }
            )
            fmt_type = workbook.add_format(
                {
                    "bold": True,
                    "text_wrap": True,
                    "valign": "top",
                    "align": "center",
                    "fg_color": "#F2F2F2",
                    "border": 1,
                    "font_size": 9,
                }
            )
            fmt_item = workbook.add_format(
                {
                    "bold": True,
                    "fg_color": "#F2F2F2",
                    "border": 1,
                    "align": "left",
                    "font_size": 10,
                }
            )
            fmt_hist = workbook.add_format(
                {
                    "fg_color": "#E8F4FD",
                    "border": 1,
                    "align": "right",
                    "num_format": "#,##0",
                }
            )
            fmt_fc = workbook.add_format(
                {
                    "fg_color": "#FFF2CC",
                    "border": 1,
                    "align": "right",
                    "num_format": "#,##0",
                }
            )

            # Write headers
            for col_num, val in enumerate(header_row_1):
                worksheet.write(0, col_num, val, fmt_month)
            for col_num, val in enumerate(header_row_2):
                worksheet.write(1, col_num, val, fmt_type)

            # Write data rows
            for row_num, data_row in enumerate(restructured_data[2:], start=2):
                for col_num, cell in enumerate(data_row):
                    if col_num == 0:
                        worksheet.write(row_num, col_num, cell, fmt_item)
                    else:
                        col_name = header_row_1[col_num] if col_num < len(header_row_1) else ""
                        try:
                            dt = pd.to_datetime(col_name, format="%b-%Y", errors="coerce")
                            if pd.isna(dt) or (dt <= max_date):
                                worksheet.write(row_num, col_num, cell, fmt_hist)
                            else:
                                worksheet.write(row_num, col_num, cell, fmt_fc)
                        except:
                            worksheet.write(row_num, col_num, cell, fmt_fc)

            # Adjust column widths
            worksheet.set_column(0, 0, 20)  # Item Code column
            for c in range(1, max_cols):
                worksheet.set_column(c, c, 12)  # Date columns
            
            # Freeze panes for better navigation
            worksheet.freeze_panes(2, 1)
    
    else:
        # Use openpyxl for basic functionality
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            restructured_df.to_excel(writer, sheet_name="MH_Sales_Forecast", index=False, header=False)
            
            # Basic formatting with openpyxl
            try:
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                
                worksheet = writer.sheets["MH_Sales_Forecast"]
                
                # Define styles
                header_font = Font(bold=True, size=10)
                header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Apply formatting to headers
                for col in range(1, max_cols + 1):
                    cell1 = worksheet.cell(row=1, column=col)
                    cell2 = worksheet.cell(row=2, column=col)
                    
                    cell1.font = header_font
                    cell1.fill = header_fill
                    cell1.border = border
                    cell1.alignment = Alignment(horizontal='center', vertical='center')
                    
                    cell2.font = header_font
                    cell2.border = border
                    cell2.alignment = Alignment(horizontal='center', vertical='center')
                
                # Set column widths
                worksheet.column_dimensions['A'].width = 20
                for col in range(2, max_cols + 1):
                    col_letter = worksheet.cell(row=1, column=col).column_letter
                    worksheet.column_dimensions[col_letter].width = 12
                    
            except ImportError:
                # If openpyxl styling is not available, just save basic Excel
                pass

    output.seek(0)
    
    # Calculate file size for user info
    file_size_mb = len(output.getvalue()) / (1024 * 1024)
    
    engine_info = "with advanced formatting" if XLSXWRITER_AVAILABLE else "basic format"
    
    st.download_button(
        label=f"📥 Download Excel ({engine_info}) - {file_size_mb:.1f}MB",
        data=output,
        file_name=f"MH_Parts_Forecast_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    
except Exception as e:
    st.error(f"❌ Error generating Excel file: {str(e)}")
    st.info("📋 You can still copy the data from the preview table above.")
    
    # Provide CSV download as fallback
    csv_output = result_df.to_csv(index=False)
    st.download_button(
        label="📥 Download as CSV (Fallback)",
        data=csv_output,
        file_name=f"MH_Parts_Forecast_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.csv",
        mime="text/csv",
    )

# Optional: Display method distribution
st.write("### 🔍 Method Distribution")
method_counts = result_df["Method"].value_counts()
st.bar_chart(method_counts)

st.write("---")
st.write("💡 **Tips for better forecasts:**")
st.write("- Ensure at least 24 months of historical data for seasonal patterns")
st.write("- Include FleetHours data to improve accuracy during seasonal periods")
st.write("- Review parts with 'Naive' methods - they may need manual adjustment")
st.write("- SARIMA models work best for parts with consistent demand patterns")
